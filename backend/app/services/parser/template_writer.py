"""批量生成 Excel 模板生成器。

`build_template_workbook(registry)` 遍历所有已注册 code_type，为每个 code_type 创建一个
sheet（名取自 `excel_sheet_name`），表头列直接从对应 schema yaml 的 `fields`（跳过 output
字段）+ `signals` 块展开。该函数是纯 I/O：不写盘，不涉及 LLM/RAG/缓存，返回内存 Workbook
对象由调用方决定如何落地（HTTP StreamingResponse 写 BytesIO，或测试里直接读字段）。

输出与 `excel_parser.parse_excel` 严格兼容（同一份 schema yaml 读列定义，确保下载→填写→
上传的闭环不会因为列位错位而失败）。
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.parser.utils import col_to_idx
from app.services.registry import CodeTypeRegistry

_HEADER_FILL = PatternFill(fgColor="D9D9D9", patternType="solid")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_PLACEHOLDER_FONT = Font(color="888888", italic=True)
_PLACEHOLDER_TEXT = "请从第 3 行开始填写实际数据（本行为说明文字，不会被解析）"


def _collect_headers(schema: dict) -> dict[int, str]:
    """返回 {1-based 列号: 表头文本}，按 schema fields(col) + signals 块展开。

    跳过 `output: true` 字段（这些是回填列，模板里不应出现）。signals 块按
    start_col / cols_per_signal / max_count / sub_fields[*].suffix 顺序展开为
    "信号N<suffix>"。
    """
    headers: dict[int, str] = {}

    for fdef in schema.get("fields", []) or []:
        if fdef.get("output"):
            continue
        col_idx = col_to_idx(fdef["col"])
        headers[col_idx] = fdef["name"]

    signals = schema.get("signals")
    if signals:
        start_idx = col_to_idx(signals["start_col"])
        cpp = int(signals["cols_per_signal"])
        max_count = int(signals["max_count"])
        sub_fields = signals.get("sub_fields", []) or []
        for i in range(max_count):
            for j, sub in enumerate(sub_fields):
                col_idx = start_idx + i * cpp + j
                headers[col_idx] = f"信号{i + 1}{sub['suffix']}"

    return headers


def build_template_workbook(registry: CodeTypeRegistry) -> openpyxl.Workbook:
    """根据 registry 中所有已注册 code_type 生成 Excel 模板 workbook。

    每个 code_type → 一个 sheet（name = ct.excel_sheet_name）。openpyxl 创建时默认会
    带一个 "Sheet"，第一次写入时复用并改名，后续 create_sheet 追加。
    """
    wb = openpyxl.Workbook()
    default_ws = wb.active
    first = True

    for ct in registry.all():
        schema = registry.get_excel_schema(ct.id)
        headers = _collect_headers(schema)

        if first:
            ws = default_ws
            ws.title = ct.excel_sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=ct.excel_sheet_name)

        for col_idx, text in headers.items():
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.fill = _HEADER_FILL

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(12, int(len(text) * 1.5))

        # 占位提示写到 B2 而不是 A2：parse_excel 在 raw_row[0] (A 列) 为空时
        # 跳过整行，这样 preflight / upload 闭环不会把说明文字误解析为一条 row_id。
        placeholder_cell = ws.cell(row=2, column=2, value=_PLACEHOLDER_TEXT)
        placeholder_cell.font = _PLACEHOLDER_FONT

    return wb
