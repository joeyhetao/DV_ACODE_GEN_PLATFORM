from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl

from app.services.registry import get_registry
from app.services.parser.utils import col_to_idx

# FEAT-18: multisheet 上传时 batch_jobs.code_type 写这个 sentinel 字面量；HTTP 端点
# 写、Celery worker 读，跨进程必须用同一字符串。任何想把 sentinel 改成别的字面量
# （例如 "multi"）只需改这一处，下游 `from .excel_parser import
# MULTISHEET_CODE_TYPE_SENTINEL` 就会自动同步。
MULTISHEET_CODE_TYPE_SENTINEL = "mixed"


@dataclass
class SignalInfo:
    name: str
    width: int
    role: str


@dataclass
class ParsedRow:
    row_id: str
    code_type: str
    intent: str
    comment: str | None = None
    # FEAT-19: 以下 6 字段在 3 列 schema 下不再由 Excel 提供，全部走 dataclass 默认值。
    # 字段名保留是为了让 batch_tasks.py 的 `row.clk` / `row.rst` / `row.signals` 等引用
    # 零改动；pipeline 6 级回退本就以这些字面量为默认值（参见 ARCHITECTURE §3.7）。
    module: str = ""
    clk: str = "clk"
    rst: str = "rst_n"
    rst_polarity: str = "低有效"
    protocol: str | None = None
    signals: list[SignalInfo] = field(default_factory=list)
    extra: dict = field(default_factory=dict)


def _parse_sheet(ws, schema: dict, code_type: str) -> list[ParsedRow]:
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    fields_def = {f["col"]: f for f in schema["fields"] if not f.get("output")}

    results: list[ParsedRow] = []
    for raw_row in rows:
        if not raw_row or not raw_row[0]:
            continue

        def get_cell(col_letter: str):
            idx = col_to_idx(col_letter) - 1
            if idx < len(raw_row):
                v = raw_row[idx]
                return str(v).strip() if v is not None else None
            return None

        extra: dict = {}
        for col, fdef in fields_def.items():
            val = get_cell(col)
            extra[fdef["field_key"]] = val

        row = ParsedRow(
            row_id=extra.get("row_id", ""),
            code_type=code_type,
            intent=extra.get("intent", "") or "",
            comment=extra.get("comment"),
            extra=extra,
        )
        results.append(row)

    return results


def parse_excel(file_path: Path, code_type: str) -> list[ParsedRow]:
    registry = get_registry()
    schema = registry.get_excel_schema(code_type)
    ct_def = registry.get(code_type)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_name = ct_def.excel_sheet_name
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        return _parse_sheet(ws, schema, code_type)
    finally:
        wb.close()


def parse_excel_multisheet(file_path: Path) -> list[ParsedRow]:
    """FEAT-18: 遍历 registry 中已注册的每种 code_type，凡 wb.sheetnames 中存在对应
    `excel_sheet_name` 的 sheet 就调 `_parse_sheet` 解析并标注 row.code_type。

    - 未知 sheet 名（不在 registry 中）静默跳过——保证 FEAT-16 多 sheet 模板兼容旧
      上传文件。
    - 全部已知 sheet 均无有效行时抛 `ValueError("no_valid_rows")`；上层 `/upload`、
      `/preflight` 端点将其转为 HTTP 400 + `detail.type="no_valid_rows"`。
    """
    registry = get_registry()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        results: list[ParsedRow] = []
        for ct in registry.all():
            if ct.excel_sheet_name not in wb.sheetnames:
                continue
            schema = registry.get_excel_schema(ct.id)
            ws = wb[ct.excel_sheet_name]
            results.extend(_parse_sheet(ws, schema, ct.id))
    finally:
        wb.close()

    if not results:
        raise ValueError("no_valid_rows")
    return results
