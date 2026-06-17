"""FEAT-18 — 批量页 sheet 自动检测多 code_type 解析路径单测。

覆盖 spec §2 / §3 全部 AC：

- `test_multisheet_only_assertion_sheet_filled` — 仅 SVA 需求 sheet 有数据行 →
  `parse_excel_multisheet` 返回行的 code_type 全为 "assertion"，行为与单 sheet
  路径等同。
- `test_multisheet_both_sheets_filled` — SVA + Coverage 两 sheet 均有数据 →
  返回行混合 "assertion" + "coverage" 两类（顺序按 registry.all() 排）。
- `test_multisheet_all_empty_raises_no_valid_rows` — 两个 sheet 均为空 →
  `parse_excel_multisheet` 抛 `ValueError("no_valid_rows")`，`POST /batch/upload`
  和 `POST /batch/preflight` 转 HTTP 400 + `detail.type == "no_valid_rows"`。
- `test_multisheet_unknown_sheet_silently_skipped` — workbook 含已知 SVA sheet
  + 未知 "Garbage" sheet（手工写入数据），multisheet 应只返回 SVA 行，Garbage
  sheet 静默跳过。
- `test_gate2_batch_reason_contains_sheet_keyword` — `_build_code_type_mismatch_reason`
  helper 输出文案含 "sheet" 关键词 + selected sheet 名 + suggested code_type，
  验证批量场景文案覆盖生效。

跑法：
    docker compose exec backend pytest tests/test_batch_multisheet_parse.py -v

测试不依赖 live PG / Redis / LLM / Qdrant：parse_excel_multisheet 是纯 I/O；
gate 2 helper 调 get_registry() 走静态 YAML；端点测试 mock 掉 DB 与
preflight_row（preflight_row 在 0 数据行时不会被调用，no_valid_rows case
也走不到那里）。
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.api.v1.batch import router as batch_router
from app.core.security import get_current_user
from app.core.database import get_db
from app.models.user import User
from app.services.parser.excel_parser import parse_excel_multisheet
from app.services.parser.template_writer import build_template_workbook
from app.services.registry import get_registry
from app.tasks.batch_tasks import _build_code_type_mismatch_reason


# ── helpers ─────────────────────────────────────────────────────────────

def _make_user() -> User:
    u = MagicMock(spec=User)
    u.id = "user-1"
    u.role = "user"
    return u


def _build_authed_app() -> FastAPI:
    """挂 batch router 的最小 FastAPI app + override get_current_user / get_db。
    avoid main.py lifespan 里的 DB / Qdrant 初始化。"""
    app = FastAPI()
    app.include_router(batch_router, prefix="/api/v1")
    app.dependency_overrides[get_current_user] = lambda: _make_user()
    # /preflight 不依赖 DB（不写 BatchJob），用一个 MagicMock 即可；但 FastAPI
    # 依赖解析仍会调用 get_db，所以 override 为 no-op async gen 兜底。
    async def _noop_db():
        yield MagicMock()
    app.dependency_overrides[get_db] = _noop_db
    return app


def _make_sva_workbook(file_path: Path, *, row_ids: list[str]) -> None:
    """生成 SVA 需求 sheet 含 row_ids 列数据行的 workbook + 空 Coverage 需求 sheet。

    FEAT-19 起 SVA schema 精简为 3 列（A row_id / B 验证意图 / C 备注），数据
    从 row 3 起仅写 A + B 两列即可命中 parse_excel_multisheet 的行识别。
    """
    wb = build_template_workbook(get_registry())
    ws = wb["SVA需求"]
    # 占位说明在 row 2（B2），数据从 row 3 起
    for i, rid in enumerate(row_ids, start=3):
        ws.cell(row=i, column=1, value=rid)  # A: row_id
        ws.cell(row=i, column=2, value=f"intent for {rid}")  # B: 验证意图
    wb.save(file_path)


def _make_dual_workbook(
    file_path: Path, *, sva_row_ids: list[str], cov_row_ids: list[str]
) -> None:
    """两 sheet 都填数据：SVA 需求 + Coverage 需求（均为 3 列 schema）。"""
    wb = build_template_workbook(get_registry())
    ws_sva = wb["SVA需求"]
    for i, rid in enumerate(sva_row_ids, start=3):
        ws_sva.cell(row=i, column=1, value=rid)
        ws_sva.cell(row=i, column=2, value=f"sva intent for {rid}")

    ws_cov = wb["Coverage需求"]
    # FEAT-19: Coverage schema 同 SVA 精简为 3 列，intent 在 B 列。
    for i, rid in enumerate(cov_row_ids, start=3):
        ws_cov.cell(row=i, column=1, value=rid)
        ws_cov.cell(row=i, column=2, value=f"cov intent for {rid}")
    wb.save(file_path)


def _make_unknown_sheet_workbook(file_path: Path, *, sva_row_ids: list[str]) -> None:
    """SVA 需求 + 一个 "Garbage" 未知 sheet（含 A 列数据）；multisheet 应只识 SVA。"""
    wb = build_template_workbook(get_registry())
    ws_sva = wb["SVA需求"]
    for i, rid in enumerate(sva_row_ids, start=3):
        ws_sva.cell(row=i, column=1, value=rid)
        ws_sva.cell(row=i, column=2, value=f"sva intent for {rid}")

    garbage = wb.create_sheet(title="Garbage")
    garbage.cell(row=1, column=1, value="编号")
    garbage.cell(row=2, column=1, value="GARBAGE-1")
    garbage.cell(row=2, column=2, value="should be ignored")
    wb.save(file_path)


# ── 1. 仅一个 sheet 有数据：行为与单 code_type 路径等同 ──────────────────

def test_multisheet_only_assertion_sheet_filled(tmp_path):
    fp = tmp_path / "sva_only.xlsx"
    _make_sva_workbook(fp, row_ids=["SVA-1", "SVA-2"])

    rows = parse_excel_multisheet(fp)

    assert len(rows) == 2, f"应得 2 条 SVA 行，实际 {len(rows)}"
    assert {r.code_type for r in rows} == {"assertion"}, (
        "所有行 code_type 应为 'assertion'，实际：" + repr([r.code_type for r in rows])
    )
    assert {r.row_id for r in rows} == {"SVA-1", "SVA-2"}


# ── 2. 多 sheet 均有数据：返回混合两类 ──────────────────────────────────

def test_multisheet_both_sheets_filled(tmp_path):
    fp = tmp_path / "dual.xlsx"
    _make_dual_workbook(
        fp, sva_row_ids=["SVA-1"], cov_row_ids=["COV-1", "COV-2"]
    )

    rows = parse_excel_multisheet(fp)

    code_types = [r.code_type for r in rows]
    assert "assertion" in code_types, f"应含 assertion 行：{code_types}"
    assert "coverage" in code_types, f"应含 coverage 行：{code_types}"
    assert len(rows) == 3, f"SVA-1 + COV-1 + COV-2 = 3 行，实际 {len(rows)}"

    by_ct = {r.code_type: [x.row_id for x in rows if x.code_type == r.code_type] for r in rows}
    assert by_ct["assertion"] == ["SVA-1"]
    assert sorted(by_ct["coverage"]) == ["COV-1", "COV-2"]


# ── 3. 全部 sheet 空 → ValueError + 端点 HTTP 400 ───────────────────────

def test_multisheet_all_empty_raises_no_valid_rows(tmp_path):
    """模板原样上传（A 列全空）触发 ValueError("no_valid_rows")。"""
    fp = tmp_path / "empty.xlsx"
    build_template_workbook(get_registry()).save(fp)

    with pytest.raises(ValueError) as exc_info:
        parse_excel_multisheet(fp)
    assert str(exc_info.value) == "no_valid_rows"


def test_preflight_no_valid_rows_returns_400_with_structured_detail():
    """空模板走 POST /batch/preflight（不传 code_type）应得 HTTP 400 +
    detail.type == "no_valid_rows"，文案含"未检测到任何有效数据行"。"""
    client = TestClient(_build_authed_app())

    buf = BytesIO()
    build_template_workbook(get_registry()).save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/v1/batch/preflight",
        files={"file": ("template.xlsx", buf.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        # 故意不传 data={"code_type": ...} —— 走 multisheet 分支
    )

    assert resp.status_code == 400, (
        f"空模板 preflight 应返回 400，实际 {resp.status_code}: {resp.text}"
    )
    body = resp.json()
    assert body["detail"]["type"] == "no_valid_rows", body
    assert "未检测到任何有效数据行" in body["detail"]["message"]


def test_preflight_corrupted_xlsx_returns_400_not_500():
    """M2: 损坏 .xlsx 文件应被翻成 HTTP 400 + 通用文案，而非以 500 暴露内部异常。"""
    client = TestClient(_build_authed_app())

    resp = client.post(
        "/api/v1/batch/preflight",
        files={"file": ("garbage.xlsx", b"not a real xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert resp.status_code == 400, (
        f"损坏 .xlsx 应返回 400，实际 {resp.status_code}: {resp.text}"
    )
    body = resp.json()
    # 不应外泄底层 openpyxl/zipfile 异常 message，只用通用提示文案。
    detail = body["detail"]
    msg = detail if isinstance(detail, str) else detail.get("message", "")
    assert "Excel" in msg or "无法读取" in msg, body


# ── 4. 未知 sheet 静默跳过 ──────────────────────────────────────────────

def test_multisheet_unknown_sheet_silently_skipped(tmp_path):
    fp = tmp_path / "with_garbage.xlsx"
    _make_unknown_sheet_workbook(fp, sva_row_ids=["SVA-1"])

    rows = parse_excel_multisheet(fp)

    assert len(rows) == 1, f"应只返回 SVA-1 一行，实际 {len(rows)}：{[r.row_id for r in rows]}"
    assert rows[0].row_id == "SVA-1"
    assert rows[0].code_type == "assertion"
    # 未知 sheet 的 GARBAGE-1 行不应出现
    assert all(r.row_id != "GARBAGE-1" for r in rows)


# ── 5. gate 2 批量文案含 "sheet" 关键词 ────────────────────────────────

def test_gate2_batch_reason_contains_sheet_keyword():
    """`_build_code_type_mismatch_reason` 输出含 "sheet" 关键词 + selected sheet
    名 + suggested code_type，验证 FEAT-18 文案覆盖生效。"""
    reason = _build_code_type_mismatch_reason(
        selected_code_type="assertion",
        suggested_code_type="coverage",
    )
    # "sheet" 关键词必须出现——这是批量场景区别于单条页文案的核心差异
    assert "sheet" in reason, f"reason 缺 'sheet' 关键词：{reason!r}"
    # selected 的 sheet 名（SVA需求）应反查到位
    assert "SVA需求" in reason, f"reason 缺 selected sheet 名 'SVA需求'：{reason!r}"
    # suggested code_type 必须明示
    assert "coverage" in reason, f"reason 缺 suggested code_type 'coverage'：{reason!r}"


def test_gate2_batch_reason_unknown_code_type_falls_back_gracefully():
    """selected_code_type 不在 registry 中时 helper 不应抛异常——退化为字面量。"""
    reason = _build_code_type_mismatch_reason(
        selected_code_type="bogus_unknown_type",
        suggested_code_type="assertion",
    )
    # registry.get 失败时退化为 selected_code_type 字面量
    assert "bogus_unknown_type" in reason
    assert "sheet" in reason


# ── 6. FEAT-19 — 3 列 schema 新场景 ────────────────────────────────────

def test_parse_3col_only_required_no_comment(tmp_path):
    """只填 A 编号 + B 意图，C 备注为空 → parse 正常返回 ParsedRow，
    comment 字段为 None 或空串，intent 字段非空。"""
    fp = tmp_path / "no_comment.xlsx"
    wb = build_template_workbook(get_registry())
    ws = wb["SVA需求"]
    ws.cell(row=3, column=1, value="ROW-1")
    ws.cell(row=3, column=2, value="check req before ack")
    # 故意不写 column=3（C 备注）
    wb.save(fp)

    rows = parse_excel_multisheet(fp)
    assert len(rows) == 1, f"应得 1 条行，实际 {len(rows)}"
    row = rows[0]
    assert row.row_id == "ROW-1"
    assert row.code_type == "assertion"
    assert row.intent == "check req before ack"
    # comment 缺失列时 _parse_sheet 的 get_cell 返回 None
    assert not row.comment, f"未填备注应为空（None 或 ''），实际 {row.comment!r}"


def test_parse_3col_empty_intent_skipped_by_worker(tmp_path):
    """A 编号有值 + B 意图空 → parse_excel_multisheet 仍返回该行（A 非空），
    但 ParsedRow.intent 为空串，batch_tasks 的 `if not intent_text` 分支
    会把这一行归为 `skipped`/`reason="空意图"`。

    这条用例锁的是 parser 侧的契约——parser 不负责跳行（A 非空就返回），
    跳行决策完全由 worker 在 batch_tasks.py:71-77 做。"""
    fp = tmp_path / "empty_intent.xlsx"
    wb = build_template_workbook(get_registry())
    ws = wb["SVA需求"]
    ws.cell(row=3, column=1, value="ROW-EMPTY")
    # column=2 (B 验证意图) 故意留空
    wb.save(fp)

    rows = parse_excel_multisheet(fp)
    assert len(rows) == 1, "A 列有值即应入列，由 worker 决定是否 skipped"
    row = rows[0]
    assert row.row_id == "ROW-EMPTY"
    assert row.intent == "", f"B 列空时 intent 应为空串，实际 {row.intent!r}"
    # 同步验证 worker 跳行条件命中：`row.intent` 假值 + `row.extra["intent"]` 也为 None
    intent_text = row.intent if row.intent else row.extra.get("intent", "")
    assert not intent_text, "worker 应将此行判为空意图"


def test_download_template_output_cols_at_d_e_f():
    """FEAT-19: 输出列 _out_template / _out_confidence / _out_status 在
    sva/coverage schema 中已前移至 D/E/F，且带 output=true 标记——模板
    下载时 template_writer 必须把 D/E/F 全部跳过，用户可见表头仅 A/B/C。"""
    wb = build_template_workbook(get_registry())
    for sheet_name in ("SVA需求", "Coverage需求"):
        ws = wb[sheet_name]
        # A/B/C 必须有表头
        assert ws.cell(row=1, column=1).value is not None, f"{sheet_name} A1 缺表头"
        assert ws.cell(row=1, column=2).value is not None, f"{sheet_name} B1 缺表头"
        assert ws.cell(row=1, column=3).value is not None, f"{sheet_name} C1 缺表头"
        # D/E/F 必须为空（output 字段不渲染到表头）
        for col_idx, letter in ((4, "D"), (5, "E"), (6, "F")):
            assert ws.cell(row=1, column=col_idx).value is None, (
                f"{sheet_name} {letter}1 不应有表头（应被 output=true 过滤），"
                f"实际 {ws.cell(row=1, column=col_idx).value!r}"
            )
