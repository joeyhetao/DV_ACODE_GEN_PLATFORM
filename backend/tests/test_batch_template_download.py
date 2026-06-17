"""FEAT-16 — 批量生成 Excel 模板下载端点单测。

覆盖 spec §2 全部 AC：

- `test_template_download_returns_valid_xlsx` — TestClient GET /batch/template
  返回 200 + 正确 Content-Type，Content-Disposition 含引号的 filename，body
  可被 openpyxl 加载。
- `test_template_has_sheet_per_code_type` — workbook.sheetnames 与
  registry.all() 的 excel_sheet_name 严格对齐。
- `test_assertion_sheet_headers_match_schema` — SVA 需求 sheet 表头列顺序
  与 sva_schema.yaml fields（跳过 output 字段）+ signals 块展开一致。
- `test_coverage_sheet_headers_match_schema` — Coverage 需求 sheet 表头列
  顺序与 coverage_schema.yaml fields（跳过 output 字段）一致，无 signals。
- `test_download_template_then_upload_roundtrip` — 生成 → BytesIO →
  excel_parser.parse_excel 返回空列表（占位行在 A 列为 None，被跳过）。
- `test_template_upload_to_preflight_returns_200` — 模板字节流走 TestClient
  上传 /batch/preflight，断言 HTTP 200 + 0 行结果（spec AC §2 第 4 条端到端
  闭环）。
- `test_template_requires_auth` — TestClient 不带 Authorization 头，验证
  OAuth2PasswordBearer 兜底 401。

跑法：
    docker compose exec backend pytest tests/test_batch_template_download.py -v

测试不依赖 live PG / Redis / LLM / Qdrant；走纯 I/O + 静态 schema yaml。
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
import yaml
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.api.v1.batch import router as batch_router
from app.core.security import get_current_user
from app.models.user import User
from app.services.parser.excel_parser import parse_excel
from app.services.parser.template_writer import build_template_workbook
from app.services.parser.utils import col_to_idx
from app.services.registry import get_registry

_SCHEMAS_DIR = Path(__file__).parent.parent / "data" / "schemas"


def _make_user() -> User:
    u = MagicMock(spec=User)
    u.id = "user-1"
    u.role = "user"
    return u


def _build_authed_app() -> FastAPI:
    """挂 batch router 的最小 FastAPI app + override get_current_user 返回 mock
    用户。避开 main.py lifespan 里的 DB / Qdrant 初始化。"""
    app = FastAPI()
    app.include_router(batch_router, prefix="/api/v1")
    app.dependency_overrides[get_current_user] = lambda: _make_user()
    return app


def _build_unauthed_app() -> FastAPI:
    """不 override，让 OAuth2PasswordBearer 在缺 token 时正常返回 401。"""
    app = FastAPI()
    app.include_router(batch_router, prefix="/api/v1")
    return app


def _expected_headers(schema_filename: str) -> dict[int, str]:
    """从 schema yaml 直接读取期望表头：跳过 output 字段。FEAT-19 起 schema
    `signals: null`，模板无 signals 块。该函数刻意复刻
    template_writer._collect_headers 的契约，用 schema yaml 真值作为表头
    断言的金标准。"""
    schema = yaml.safe_load((_SCHEMAS_DIR / schema_filename).read_text(encoding="utf-8"))
    out: dict[int, str] = {}
    for fdef in schema.get("fields", []) or []:
        if fdef.get("output"):
            continue
        out[col_to_idx(fdef["col"])] = fdef["name"]
    return out


# ── 1. 端到端：TestClient 拉模板，断言 HTTP 元数据 + body 可解析 ──────────

def test_template_download_returns_valid_xlsx():
    client = TestClient(_build_authed_app())
    resp = client.get("/api/v1/batch/template")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = resp.headers["content-disposition"]
    today = date.today().strftime("%Y%m%d")
    # filename 走 RFC 6266 quoted-string 形式
    assert "attachment;" in cd
    assert f'filename="batch_template_{today}.xlsx"' in cd

    assert len(resp.content) > 0
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert len(wb.sheetnames) >= 1


# ── 2. sheetnames 与 registry 严格对齐 ──────────────────────────────────

def test_template_has_sheet_per_code_type():
    registry = get_registry()
    wb = build_template_workbook(registry)
    expected = {ct.excel_sheet_name for ct in registry.all()}
    assert set(wb.sheetnames) == expected
    # 至少包含 spec 明确点名的 2 个 sheet
    assert "SVA需求" in wb.sheetnames
    assert "Coverage需求" in wb.sheetnames


# ── 3. SVA 表头匹配 sva_schema.yaml ────────────────────────────────────

def test_assertion_sheet_headers_match_schema():
    wb = build_template_workbook(get_registry())
    ws = wb["SVA需求"]
    expected = _expected_headers("sva_schema.yaml")

    for col_idx, text in expected.items():
        assert ws.cell(row=1, column=col_idx).value == text, (
            f"SVA sheet col={col_idx} 表头不匹配：期望 {text!r}, "
            f"实际 {ws.cell(row=1, column=col_idx).value!r}"
        )

    # FEAT-19: schema 精简为 3 列（A 编号 / B 验证意图 / C 备注），
    # signals 块已删除——所有 module/clk/rst/signals 等旧表头不应再出现。
    assert set(expected.values()) == {"编号", "验证意图", "备注"}, (
        f"SVA sheet 用户可见表头应为 3 列固定集合，实际：{expected.values()!r}"
    )


# ── 4. Coverage 表头匹配 coverage_schema.yaml ──────────────────────────

def test_coverage_sheet_headers_match_schema():
    wb = build_template_workbook(get_registry())
    ws = wb["Coverage需求"]
    expected = _expected_headers("coverage_schema.yaml")

    for col_idx, text in expected.items():
        assert ws.cell(row=1, column=col_idx).value == text, (
            f"Coverage sheet col={col_idx} 表头不匹配：期望 {text!r}, "
            f"实际 {ws.cell(row=1, column=col_idx).value!r}"
        )

    # FEAT-19: Coverage schema 同步精简为 3 列（A 编号 / B 覆盖意图 / C 备注），
    # main_signal_* / cross*_* / bin_hint / sample_condition 等旧字段全部移除。
    assert set(expected.values()) == {"编号", "覆盖意图", "备注"}, (
        f"Coverage sheet 用户可见表头应为 3 列固定集合，实际：{expected.values()!r}"
    )


# ── 5. 闭环：生成 → 落盘 → excel_parser 解析为空列表 ────────────────────

def test_download_template_then_upload_roundtrip(tmp_path):
    wb = build_template_workbook(get_registry())

    for code_type, sheet_name in (("assertion", "SVA需求"), ("coverage", "Coverage需求")):
        path = tmp_path / f"{code_type}.xlsx"
        wb.save(path)
        rows = parse_excel(path, code_type)
        # 占位提示写在 B2，A2 (row_id 列) 为 None，parse_excel 会跳过整行。
        # 这是 spec AC §2 第 4 条「空行不报错」的关键不变量：模板原样上传应
        # 产生 0 条 ParsedRow。
        assert rows == [], (
            f"{code_type} 模板原样解析应得 0 行，实际 {len(rows)} 行：{rows!r}"
        )


# ── 6. 端到端闭环：模板上传 /preflight 返回 200 + 空结果 ────────────────

def test_template_upload_to_preflight_returns_200():
    """spec AC §2 第 4 条「空行不报错」端到端：把 build_template_workbook 输出
    的字节流原样上传到 /batch/preflight，验证 HTTP 200 + 0 行结果。

    覆盖范围警告：此测试仅验证「0 行短路路径」。因为模板里 A 列没有数据
    （占位提示在 B2），parse_excel 返回空列表，于是 batch.py 的 preflight
    handler 中 `for row in rows` 循环体从不执行，`preflight_row` **从不被
    调用**。这意味着测试 pass 不代表 preflight 调用链 + Qdrant/embedding
    依赖被验证。

    如果将来要测试模板里有数据行的 preflight 路径（preflight_row 被调用、
    走到 RAG），必须 mock preflight_row 或起完整 Qdrant/embedding 桩。
    本测试在模板永远是 0 数据行的合同下是健康的，不需要 mock。"""
    client = TestClient(_build_authed_app())

    # 把 workbook 写入 BytesIO 模拟上传文件
    buf = BytesIO()
    build_template_workbook(get_registry()).save(buf)
    buf.seek(0)

    for code_type in ("assertion", "coverage"):
        buf.seek(0)
        resp = client.post(
            "/api/v1/batch/preflight",
            files={"file": ("template.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"code_type": code_type},
        )
        assert resp.status_code == 200, (
            f"preflight {code_type} 应返回 200，实际 {resp.status_code}: {resp.text}"
        )
        body = resp.json()
        assert body.get("results") == [], (
            f"模板原样上传 preflight 应产生 0 行结果，实际：{body!r}"
        )


# ── 7. 无 Bearer token → 401（OAuth2PasswordBearer 兜底） ──────────────

def test_template_requires_auth():
    # 不 override get_current_user：OAuth2PasswordBearer 在请求头缺 token 时
    # 会在依赖解析的最外层抛 401，所以 get_db 永远不会被调用，无需 DB 桩。
    client = TestClient(_build_unauthed_app())
    resp = client.get("/api/v1/batch/template")
    assert resp.status_code == 401
