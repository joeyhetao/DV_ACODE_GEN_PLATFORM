"""
Generate SVA and Coverage requirement Excel template files.
Run: python gen_templates.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ─────────── Color palette ───────────
CLR_HEADER_BG   = "1F4E79"   # dark blue
CLR_HEADER_FG   = "FFFFFF"
CLR_INPUT_BG    = "FFFFFF"
CLR_INPUT_EVEN  = "F2F7FF"   # light blue stripe
CLR_OUTPUT_BG   = "D9D9D9"   # gray – read-only output
CLR_OUTPUT_HDR  = "595959"   # dark gray header for output cols
CLR_REQUIRED_BG = "FFF2CC"   # yellow – required field hint row
CLR_SECTION_BG  = "BDD7EE"   # group separator header

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_font(is_output=False):
    return Font(
        bold=True,
        color=CLR_HEADER_FG if not is_output else CLR_HEADER_FG,
        size=10,
        name="微软雅黑"
    )


def cell_font(bold=False):
    return Font(bold=bold, size=10, name="微软雅黑")


def fill(color):
    return PatternFill("solid", fgColor=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_cell(cell, is_output=False):
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="微软雅黑")
    cell.fill = fill(CLR_OUTPUT_HDR if is_output else CLR_HEADER_BG)
    cell.alignment = center()
    cell.border = BORDER


def style_data_cell(cell, row_even=False, is_output=False):
    if is_output:
        cell.fill = fill(CLR_OUTPUT_BG)
    else:
        cell.fill = fill(CLR_INPUT_EVEN if row_even else CLR_INPUT_BG)
    cell.font = cell_font()
    cell.alignment = left()
    cell.border = BORDER


def add_dv(ws, formula_or_list, col_letter, first_row, last_row, is_list=True):
    """Add data validation to a column range."""
    if is_list:
        dv = DataValidation(
            type="list",
            formula1=formula_or_list,
            allow_blank=True,
            showDropDown=False,
        )
    else:
        dv = DataValidation(
            type="custom",
            formula1=formula_or_list,
            allow_blank=True,
        )
    dv.sqref = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


# ═══════════════════════════════════════════════
#  SVA 断言需求表
# ═══════════════════════════════════════════════
def build_sva(wb):
    ws = wb.create_sheet("SVA断言需求表")
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A3"     # freeze first 2 header rows

    # ── Column definitions ──
    # (col_letter, width, display_name, sub_header, is_output)
    columns = [
        # ─ 模块基本信息 ─
        ("A",  12, "模块名称",       "module_name",        False),
        ("B",  12, "子模块/接口",    "submodule",          False),
        ("C",  10, "协议类型",       "protocol",           False),
        # ─ 时钟复位 ─
        ("D",  12, "时钟信号名",     "clk_signal",         False),
        ("E",  12, "复位信号名",     "rst_signal",         False),
        ("F",  10, "复位极性",       "rst_polarity",       False),
        # ─ 信号列表（最多 8 个信号）─
        ("G",  14, "信号1名称",      "sig1_name",          False),
        ("H",  10, "信号1位宽",      "sig1_width",         False),
        ("I",  12, "信号1角色",      "sig1_role",          False),
        ("J",  14, "信号2名称",      "sig2_name",          False),
        ("K",  10, "信号2位宽",      "sig2_width",         False),
        ("L",  12, "信号2角色",      "sig2_role",          False),
        ("M",  14, "信号3名称",      "sig3_name",          False),
        ("N",  10, "信号3位宽",      "sig3_width",         False),
        ("O",  12, "信号3角色",      "sig3_role",          False),
        ("P",  14, "信号4名称",      "sig4_name",          False),
        ("Q",  10, "信号4位宽",      "sig4_width",         False),
        ("R",  12, "信号4角色",      "sig4_role",          False),
        # ─ 验证意图 ─
        ("S",  40, "验证意图（自然语言描述）", "intent",   False),
        # ─ 附加约束 ─
        ("T",  14, "最大时钟周期数", "max_cycles",         False),
        ("U",  12, "严重级别",       "severity",           False),
        # ─ 输出列（系统填写，灰色）─
        ("V",  16, "匹配模板ID",     "[系统输出]",         True),
        ("W",  10, "置信度",         "[系统输出]",         True),
        ("X",  12, "生成状态",       "[系统输出]",         True),
    ]

    # ── Row 1: group headers ──
    groups = [
        (1, 3,  "模块基本信息",   CLR_SECTION_BG,  "1F4E79"),
        (4, 6,  "时钟与复位",     "D6E4F0",        "1F4E79"),
        (7, 18, "信号列表（最多4组，可扩展）", CLR_SECTION_BG, "1F4E79"),
        (19,19, "验证意图",       "FFE699",        "7F6000"),
        (20,21, "附加约束",       "D6E4F0",        "1F4E79"),
        (22,24, "系统输出（勿填）",CLR_OUTPUT_BG,  "595959"),
    ]
    for start_col, end_col, label, bg, fg in groups:
        sc = get_column_letter(start_col)
        ec = get_column_letter(end_col)
        if start_col != end_col:
            ws.merge_cells(f"{sc}1:{ec}1")
        cell = ws[f"{sc}1"]
        cell.value = label
        cell.font = Font(bold=True, color=fg, size=10, name="微软雅黑")
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = BORDER

    # ── Row 2: column headers ──
    for idx, (col, width, name, sub, is_out) in enumerate(columns, start=1):
        ws.column_dimensions[col].width = width
        c2 = ws.cell(row=2, column=idx, value=name)
        style_header_cell(c2, is_output=is_out)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    # ── Data rows (examples) ──
    LAST_DATA_ROW = 102   # pre-fill 100 rows
    for r in range(3, LAST_DATA_ROW + 1):
        even = (r % 2 == 0)
        for idx, (col, _, _, _, is_out) in enumerate(columns, start=1):
            c = ws.cell(row=r, column=idx)
            style_data_cell(c, row_even=even, is_output=is_out)
        ws.row_dimensions[r].height = 18

    # ── Example rows ──
    examples = [
        # AXI4 握手数据稳定性
        ["dma_top", "AXI4 Write Channel", "AXI4",
         "clk", "rst_n", "低有效",
         "awvalid", "1", "valid",
         "awready", "1", "ready",
         "awaddr",  "32","data",
         "awlen",   "8", "count",
         "当awvalid拉高后，在awready到来之前，awaddr和awlen必须保持稳定",
         "10", "error",
         "", "", ""],
        # AXI4 AW通道响应超时
        ["dma_top", "AXI4 Write Channel", "AXI4",
         "clk", "rst_n", "低有效",
         "awvalid", "1", "valid",
         "awready", "1", "ready",
         "", "", "",
         "", "", "",
         "awvalid拉高后，awready必须在最多16个周期内响应",
         "16", "error",
         "", "", ""],
        # APB PSEL/PENABLE握手
        ["apb_slave", "APB Interface", "APB",
         "pclk", "presetn", "低有效",
         "psel",    "1", "enable",
         "penable", "1", "enable",
         "pready",  "1", "ready",
         "prdata",  "32","data",
         "APB访问时PENABLE必须在PSEL拉高后的下一个周期拉高，PREADY低时数据保持稳定",
         "1", "error",
         "", "", ""],
    ]

    for row_idx, ex in enumerate(examples, start=3):
        even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(ex, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            is_out = col_idx >= 22
            style_data_cell(c, row_even=even, is_output=is_out)

    # ── Dropdown validations ──
    POLARITY = '"低有效,高有效"'
    PROTOCOL = '"AXI4,AXI4-Lite,AXI3,AHB,APB,自定义"'
    ROLE     = '"valid,ready,data,state,req,ack,start,end,enable,count,other"'
    SEVERITY = '"error,warning,info"'

    add_dv(ws, POLARITY, "F", 3, LAST_DATA_ROW)
    add_dv(ws, PROTOCOL, "C", 3, LAST_DATA_ROW)
    for col in ["I", "L", "O", "R"]:
        add_dv(ws, ROLE, col, 3, LAST_DATA_ROW)
    add_dv(ws, SEVERITY, "U", 3, LAST_DATA_ROW)

    # ── Tab color ──
    ws.sheet_properties.tabColor = "1F4E79"

    return ws


# ═══════════════════════════════════════════════
#  功能覆盖率需求表
# ═══════════════════════════════════════════════
def build_coverage(wb):
    ws = wb.create_sheet("功能覆盖率需求表")
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A3"

    columns = [
        # ─ 模块基本信息 ─
        ("A",  12, "模块名称",         "module_name",      False),
        ("B",  12, "子模块/接口",      "submodule",        False),
        ("C",  10, "协议类型",         "protocol",         False),
        # ─ 时钟复位 ─
        ("D",  12, "时钟信号名",       "clk_signal",       False),
        ("E",  12, "复位信号名",       "rst_signal",       False),
        ("F",  10, "复位极性",         "rst_polarity",     False),
        # ─ 覆盖对象 ─
        ("G",  14, "主信号名",         "primary_signal",   False),
        ("H",  10, "主信号位宽",       "primary_width",    False),
        ("I",  10, "数据类型",         "data_type",        False),
        # ─ 覆盖类型 ─
        ("J",  14, "覆盖类型",         "coverage_type",    False),
        # ─ 交叉覆盖信号 ─
        ("K",  14, "交叉信号1名称",    "cross1_name",      False),
        ("L",  10, "交叉信号1位宽",    "cross1_width",     False),
        ("M",  14, "交叉信号2名称",    "cross2_name",      False),
        ("N",  10, "交叉信号2位宽",    "cross2_width",     False),
        # ─ 值域约束 ─
        ("O",  14, "最小值/枚举列表",  "min_or_enum",      False),
        ("P",  14, "最大值/步长",      "max_or_step",      False),
        ("Q",  16, "非法值/排除值",    "illegal_values",   False),
        # ─ 验证意图 ─
        ("R",  40, "验证意图（自然语言描述）","intent",    False),
        # ─ 权重 ─
        ("S",  10, "采样权重",         "weight",           False),
        # ─ 输出列 ─
        ("T",  16, "匹配模板ID",       "[系统输出]",       True),
        ("U",  10, "置信度",           "[系统输出]",       True),
        ("V",  12, "生成状态",         "[系统输出]",       True),
    ]

    groups = [
        (1,  3,  "模块基本信息",        CLR_SECTION_BG, "1F4E79"),
        (4,  6,  "时钟与复位",          "D6E4F0",       "1F4E79"),
        (7,  10, "覆盖对象与类型",      CLR_SECTION_BG, "1F4E79"),
        (11, 14, "交叉覆盖信号（可选）","D6E4F0",       "1F4E79"),
        (15, 17, "值域约束（可选）",    CLR_SECTION_BG, "1F4E79"),
        (18, 18, "验证意图",            "FFE699",       "7F6000"),
        (19, 19, "采样权重",            "D6E4F0",       "1F4E79"),
        (20, 22, "系统输出（勿填）",    CLR_OUTPUT_BG,  "595959"),
    ]

    for start_col, end_col, label, bg, fg in groups:
        sc = get_column_letter(start_col)
        ec = get_column_letter(end_col)
        if start_col != end_col:
            ws.merge_cells(f"{sc}1:{ec}1")
        cell = ws[f"{sc}1"]
        cell.value = label
        cell.font = Font(bold=True, color=fg, size=10, name="微软雅黑")
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = BORDER

    for idx, (col, width, name, sub, is_out) in enumerate(columns, start=1):
        ws.column_dimensions[col].width = width
        c2 = ws.cell(row=2, column=idx, value=name)
        style_header_cell(c2, is_output=is_out)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    LAST_DATA_ROW = 102
    for r in range(3, LAST_DATA_ROW + 1):
        even = (r % 2 == 0)
        for idx, (col, _, _, _, is_out) in enumerate(columns, start=1):
            c = ws.cell(row=r, column=idx)
            style_data_cell(c, row_even=even, is_output=is_out)
        ws.row_dimensions[r].height = 18

    examples = [
        # AXI4 突发长度枚举覆盖
        ["dma_top", "AXI4 Write Channel", "AXI4",
         "clk", "rst_n", "低有效",
         "awlen", "8", "无符号整数",
         "值覆盖",
         "", "", "", "",
         "0,1,3,7,15,255", "", "128,129",
         "覆盖AXI4突发传输长度的典型值：单拍(0)、2拍(1)、4拍(3)、8拍(7)、16拍(15)及最大256拍(255)",
         "1",
         "", "", ""],
        # AXI4 写操作类型 × 突发长度交叉覆盖
        ["dma_top", "AXI4 Write Channel", "AXI4",
         "clk", "rst_n", "低有效",
         "awburst", "2", "枚举",
         "交叉覆盖",
         "awlen", "8", "", "",
         "0,1,2", "", "",
         "交叉覆盖突发类型（FIXED/INCR/WRAP）与突发长度（0/7/15/255）的所有合法组合",
         "1",
         "", "", ""],
        # FSM 状态转移覆盖
        ["arbiter", "状态机", "自定义",
         "clk", "rst_n", "低有效",
         "state", "3", "枚举",
         "状态转移覆盖",
         "", "", "", "",
         "IDLE,REQ,GRANT,BUSY,DONE", "", "",
         "覆盖仲裁器FSM所有合法状态转移路径，重点覆盖IDLE→REQ→GRANT和GRANT→BUSY→DONE序列",
         "2",
         "", "", ""],
    ]

    for row_idx, ex in enumerate(examples, start=3):
        even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(ex, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            is_out = col_idx >= 20
            style_data_cell(c, row_even=even, is_output=is_out)

    POLARITY  = '"低有效,高有效"'
    PROTOCOL  = '"AXI4,AXI4-Lite,AXI3,AHB,APB,自定义"'
    DATA_TYPE = '"无符号整数,有符号整数,枚举,布尔,浮点"'
    COV_TYPE  = '"值覆盖,状态转移覆盖,交叉覆盖,协议事务覆盖,异常场景覆盖"'

    add_dv(ws, POLARITY,  "F", 3, LAST_DATA_ROW)
    add_dv(ws, PROTOCOL,  "C", 3, LAST_DATA_ROW)
    add_dv(ws, DATA_TYPE, "I", 3, LAST_DATA_ROW)
    add_dv(ws, COV_TYPE,  "J", 3, LAST_DATA_ROW)

    ws.sheet_properties.tabColor = "375623"

    return ws


# ═══════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ─ SVA workbook ─
wb_sva = Workbook()
wb_sva.remove(wb_sva.active)   # remove default Sheet
build_sva(wb_sva)
sva_path = os.path.join(OUTPUT_DIR, "sva_requirements_template.xlsx")
wb_sva.save(sva_path)
print(f"[OK] {sva_path}")

# ─ Coverage workbook ─
wb_cov = Workbook()
wb_cov.remove(wb_cov.active)
build_coverage(wb_cov)
cov_path = os.path.join(OUTPUT_DIR, "coverage_requirements_template.xlsx")
wb_cov.save(cov_path)
print(f"[OK] {cov_path}")
